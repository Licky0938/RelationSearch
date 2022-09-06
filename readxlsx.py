import openpyxl

class readxlsx:
    def __init__(self, bookname: str, sheetnum: int) -> None:
        self.wb = openpyxl.load_workbook(bookname)
        self.ws = self.wb.worksheets[sheetnum]
        print("[System]", bookname, "read")

    # private functions
    def __get_rowrange(self, col: int) -> int:
        row = 1
        keep = 1050000
        indata = True
        while True:
            sts = self.ws.cell(row=row, column=col).value is None
            # print("keep:", keep, "row:", row, "indata:", indata, "sts:", sts)
            
            if indata ^ sts:
                if keep > row:
                    keep = row
                else:
                    return keep, row - 1

            indata = sts
            row += 1
    
    def __getgen_bycol(self, s_col: int, e_col:int) -> tuple:
        s_row, e_row = self.__get_rowrange(s_col)
        return self.ws.iter_rows(min_row=s_row, max_row=e_row, min_col=s_col, max_col=e_col)
    
    # public functions
    def get_targetset(self, start: int, end: int = None) -> set:
        if end is None:
            end = start
        t_md = self.__getgen_bycol(start, end)
        s_md = set()
        for row in t_md:
            for cell in row:
                s_md.add(cell.value)
        return s_md

    def get_valuelist_bycol(self, s_col: int, e_col:int) -> list:
        t_2d = self.__getgen_bycol(s_col, e_col)
        return ([[cell.value for cell in row] for row in t_2d])
    
    def gen_linkdict(self, l_2d: list, target:set = None) -> dict:
        link = {}
        if len(l_2d[0]) < 2:
            print("[Error]", self, "does not match size")
            return None
        if target is None:
            target = set(x[0] for x in l_2d)
        for l_val in l_2d:
            if len(set(l_val) & target) > 0:
                if l_val[0] not in link:
                    link[l_val[0]] = []
                for i in range(1, len(l_val)):
                    link[l_val[0]].append(l_val[i])

        print("[System]", "link dictionary generated")
        return link
    
    def get_link(self, s_col: int, e_col: int, target: set = None) -> dict:
        return self.gen_linkdict(self.get_valuelist_bycol(s_col, e_col), target)



if __name__ == "__main__":
    
    FILENAME = "jobconnection.xlsx"
    x = readxlsx(FILENAME, "Sheet1")
    a = x.get_link(1, 2, ["I"])
    b = x.get_valuelist_bycol(1, 2)
    # c = x.get_targetdict(1)
    print(a)